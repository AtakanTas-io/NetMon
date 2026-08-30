import io
import time

import pytest
import server
from conftest import persistent_test_client
from openpyxl import load_workbook


@pytest.fixture()
def isolated_server(tmp_path, monkeypatch, test_portal):
    db_path = tmp_path / "netmon-phase4.db"
    password_path = tmp_path / "initial-admin.txt"
    monkeypatch.setattr(server, "DB_PATH", db_path)
    monkeypatch.setattr(server, "INITIAL_PASSWORD_PATH", password_path)
    server._devices_cache.update({"ts": 0, "data": [], "error": None, "scan_status": "idle"})
    server._api_key_rate_state.clear()
    server.init_db()
    with persistent_test_client(server.app, test_portal) as client:
        yield client, password_path


def _bootstrap_admin(client, password_path):
    initial_password = password_path.read_text(encoding="utf-8").splitlines()[1]
    login = client.post("/api/auth/login", json={"username": "admin", "password": initial_password})
    headers = {"Authorization": f"Bearer {login.json()['token']}"}
    changed = client.post(
        "/api/auth/change-password",
        headers=headers,
        json={"current_password": initial_password, "new_password": "Phase4-Strong-Pass-2026!"},
    )
    assert changed.status_code == 200
    return headers


def test_phase4_schema_is_migrated(isolated_server):
    with server.db_conn() as conn:
        tables = {row[0] for row in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()}
        asset_columns = {row[1] for row in conn.execute("PRAGMA table_info(inventory_assets)")}

    assert {
        "sites",
        "operational_snapshots",
        "alert_rules",
        "alert_events",
        "alert_user_states",
        "report_schedules",
        "report_runs",
        "api_keys",
    }.issubset(tables)
    assert "site_id" in asset_columns


def test_alarm_inbox_read_and_suppressed_state_is_persistent(isolated_server):
    client, password_path = isolated_server
    headers = _bootstrap_admin(client, password_path)
    alert_ts = round(time.time(), 6)
    with server.db_conn() as conn:
        conn.execute(
            "INSERT INTO alerts(ts,level,message,source) VALUES(?,?,?,?)",
            (alert_ts, "critical", "10.0.0.9 cihazı çevrimdışı", "test"),
        )
        conn.commit()

    inbox = client.get("/api/alerts/inbox", headers=headers).json()
    assert inbox["unread"] == 1
    alert_id = inbox["alerts"][0]["id"]
    changed = client.put(f"/api/alerts/{alert_id}/state", headers=headers, json={"is_read": True, "suppressed": True})
    assert changed.status_code == 200
    persisted = client.get("/api/alerts/inbox", headers=headers).json()
    assert persisted["unread"] == 0
    assert persisted["alerts"][0]["is_read"] is True
    assert persisted["alerts"][0]["suppressed"] is True


def test_user_theme_preference_persists_without_admin_settings_permission(isolated_server):
    client, password_path = isolated_server
    headers = _bootstrap_admin(client, password_path)
    saved = client.put("/api/preferences", headers=headers, json={"theme": "light"})
    assert saved.status_code == 200
    assert client.get("/api/preferences", headers=headers).json()["theme"] == "light"
    assert client.put("/api/preferences", headers=headers, json={"theme": "invalid"}).status_code == 422


def test_alert_rule_uses_evidence_and_respects_cooldown(isolated_server):
    client, password_path = isolated_server
    headers = _bootstrap_admin(client, password_path)
    now = time.time()
    with server.db_conn() as conn:
        conn.execute(
            "INSERT INTO known_devices(mac,hostname,first_seen,last_seen,last_ip,last_status) VALUES(?,?,?,?,?,?)",
            ("AA:BB:CC:DD:EE:01", "old-switch", now - 7200, now - 3600, "10.0.0.2", "offline"),
        )
        conn.commit()

    created = client.post(
        "/api/alert-rules",
        headers=headers,
        json={
            "name": "Uzun süre çevrimdışı",
            "rule_type": "offline_duration",
            "threshold_seconds": 1800,
            "level": "critical",
            "cooldown_seconds": 3600,
        },
    )
    assert created.status_code == 200
    first = client.post("/api/alert-rules/evaluate", headers=headers)
    second = client.post("/api/alert-rules/evaluate", headers=headers)

    assert first.status_code == 200
    assert len(first.json()["events"]) == 1
    assert first.json()["events"][0]["evidence"][0]["ip"] == "10.0.0.2"
    assert second.json()["events"] == []
    events = client.get("/api/alert-events", headers=headers).json()["events"]
    assert len(events) == 1
    assert events[0]["delivery"] == {}


def test_ip_conflict_rule_uses_current_discovery_cache(isolated_server):
    client, password_path = isolated_server
    headers = _bootstrap_admin(client, password_path)
    server._devices_cache["data"] = [
        {"ip": "10.0.0.8", "mac": "AA:AA:AA:AA:AA:01"},
        {"ip": "10.0.0.8", "mac": "AA:AA:AA:AA:AA:02"},
    ]
    client.post(
        "/api/alert-rules",
        headers=headers,
        json={"name": "IP çakışması", "rule_type": "ip_conflict", "cooldown_seconds": 60},
    )

    result = client.post("/api/alert-rules/evaluate", headers=headers)

    assert result.status_code == 200
    assert result.json()["events"][0]["evidence"] == [
        {"ip": "10.0.0.8", "macs": ["AA:AA:AA:AA:AA:01", "AA:AA:AA:AA:AA:02"]}
    ]


def test_history_ranges_and_measured_snapshot(isolated_server):
    client, password_path = isolated_server
    headers = _bootstrap_admin(client, password_path)
    now = time.time()
    with server.db_conn() as conn:
        conn.execute(
            "INSERT INTO traffic(ts,wifi_sent,wifi_recv,eth_sent,eth_recv) VALUES(?,?,?,?,?)", (now, 1, 2, 3, 4)
        )
        conn.execute(
            "INSERT INTO operational_snapshots(ts,device_count,online_count,open_port_count,traffic_bps,source) "
            "VALUES(?,?,?,?,?,?)",
            (now - 2 * 86400, 1, 1, 1, 4, "measured"),
        )
        conn.commit()
    server._devices_cache["data"] = [
        {"ip": "10.0.0.10", "status": "online", "classification": {"open_ports": [22, 443]}}
    ]

    snapshot = client.post("/api/history/snapshot", headers=headers)
    last_day = client.get("/api/history?range=24h", headers=headers).json()["points"]
    last_week = client.get("/api/history?range=7d", headers=headers).json()["points"]

    assert snapshot.json()["inserted"] == 1
    assert len(last_day) == 1
    assert last_day[0]["traffic_bps"] == 10
    assert last_day[0]["source"] == "measured_and_discovered"
    assert len(last_week) == 2


def test_sites_assign_assets_by_private_cidr(isolated_server):
    client, password_path = isolated_server
    headers = _bootstrap_admin(client, password_path)
    now = time.time()
    with server.db_conn() as conn:
        conn.execute(
            "INSERT INTO inventory_assets(identity_key,hostname,ip_address,status,first_seen,last_seen,created_at,updated_at) "
            "VALUES(?,?,?,?,?,?,?,?)",
            ("mac:phase4", "branch-pc", "10.25.1.9", "online", now, now, now, now),
        )
        conn.commit()

    invalid = client.post("/api/sites", headers=headers, json={"name": "Genel", "cidrs": ["8.8.8.0/24"]})
    created = client.post(
        "/api/sites",
        headers=headers,
        json={"name": "İstanbul", "description": "Merkez", "cidrs": ["10.25.0.0/16"]},
    )
    overlap = client.post("/api/sites", headers=headers, json={"name": "Çakışan", "cidrs": ["10.25.1.0/24"]})
    sites = client.get("/api/sites", headers=headers).json()["sites"]

    assert invalid.status_code == 400
    assert created.status_code == 200
    assert overlap.status_code == 409
    assert sites[0]["asset_count"] == 1
    with server.db_conn() as conn:
        assert (
            conn.execute("SELECT site_id FROM inventory_assets WHERE identity_key='mac:phase4'").fetchone()[0]
            == created.json()["id"]
        )


def test_pdf_and_excel_exports_are_valid_files(isolated_server):
    client, password_path = isolated_server
    headers = _bootstrap_admin(client, password_path)

    pdf = client.get("/api/reports/export?format=pdf", headers=headers)
    xlsx = client.get("/api/reports/export?format=xlsx", headers=headers)

    assert pdf.status_code == 200
    assert pdf.content.startswith(b"%PDF-")
    assert pdf.headers["x-netmon-data-source"] == "stored-observations"
    workbook = load_workbook(io.BytesIO(xlsx.content), read_only=True)
    assert workbook.sheetnames == ["Envanter", "Geçmiş"]
    assert workbook["Envanter"]["I1"].value == "Veri Kaynağı"


def test_due_schedule_creates_a_report_run_without_fake_email(isolated_server):
    client, password_path = isolated_server
    headers = _bootstrap_admin(client, password_path)
    created = client.post(
        "/api/report-schedules",
        headers=headers,
        json={"name": "Saatlik rapor", "format": "xlsx", "interval_seconds": 3600},
    )
    schedule_id = created.json()["id"]
    with server.db_conn() as conn:
        conn.execute("UPDATE report_schedules SET next_run_at=? WHERE id=?", (time.time() - 1, schedule_id))
        conn.commit()

    run = client.post("/api/report-schedules/run-due", headers=headers)

    assert run.json()["processed"] == 1
    with server.db_conn() as conn:
        assert (
            conn.execute("SELECT status FROM report_runs WHERE schedule_id=?", (schedule_id,)).fetchone()[0]
            == "generated"
        )
    runs = client.get("/api/report-runs", headers=headers).json()["runs"]
    assert runs[0]["schedule_id"] == schedule_id
    updated = client.put(
        f"/api/report-schedules/{schedule_id}",
        headers=headers,
        json={"name": "Günlük rapor", "format": "pdf", "interval_seconds": 86400},
    )
    assert updated.status_code == 200
    assert client.delete(f"/api/report-schedules/{schedule_id}", headers=headers).status_code == 200


def test_api_key_is_shown_once_limited_and_revocable(isolated_server):
    client, password_path = isolated_server
    headers = _bootstrap_admin(client, password_path)
    created = client.post(
        "/api/api-keys",
        headers=headers,
        json={"name": "Rapor otomasyonu", "permissions": ["reports.view"], "rate_limit_per_minute": 5},
    )
    body = created.json()
    key_headers = {"Authorization": f"Bearer {body['key']}"}

    assert created.status_code == 200
    assert client.get("/api/history", headers=key_headers).status_code == 200
    assert client.get("/api/sites", headers=key_headers).status_code == 403
    listed = client.get("/api/api-keys", headers=headers).json()["keys"]
    assert "key" not in listed[0]

    # İki önceki API anahtarı isteği de hız sayacına dahildir.
    assert client.get("/api/history", headers=key_headers).status_code == 200
    assert client.get("/api/history", headers=key_headers).status_code == 200
    assert client.get("/api/history", headers=key_headers).status_code == 200
    assert client.get("/api/history", headers=key_headers).status_code == 429

    revoked = client.delete(f"/api/api-keys/{body['id']}", headers=headers)
    assert revoked.status_code == 200
    assert client.get("/api/history", headers=key_headers).status_code == 401
